"""
MythosCards Exporter - I/O Operations (v2.0)
Excel dosya okuma ve yazma işlemleri

DEĞİŞİKLİK: Çıktı sheet'inde yapılandırılmış kolonlar
- Kart Listesi (eski format, geriye uyumluluk)
- player_name
- series_name  
- group
- denominator
- is_signed
- Görsel Dosyası (Part 2 için)
"""

import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from datetime import datetime

from utils import (
    FileOperationError, handle_file_collision, 
    ist_timestamp, get_current_user, safe_filename
)
from version import PROGRAM_VERSION, PROGRAM_NAME

logger = logging.getLogger(__name__)


class ExcelReader:
    """Excel dosya okuma sınıfı"""
    
    def __init__(self, file_path: Path):
        self.file_path = Path(file_path)
        self.workbook = None
        self.data = None
        
    def read_checklist(self) -> pd.DataFrame:
        """Checklist Excel dosyasını oku"""
        try:
            if not self.file_path.exists():
                raise FileOperationError(f"Excel dosyası bulunamadı: {self.file_path}")
            
            logger.info(f"Excel dosyası okunuyor: {self.file_path}")
            
            # Openpyxl ile oku (daha iyi Unicode desteği için)
            self.workbook = load_workbook(self.file_path, data_only=True)
            
            # İlk sheet'i al veya 'Checklist' isimli sheet'i ara
            sheet_name = None
            if 'Checklist' in self.workbook.sheetnames:
                sheet_name = 'Checklist'
            else:
                sheet_name = self.workbook.sheetnames[0]
            
            # Pandas ile oku
            self.data = pd.read_excel(
                self.file_path, 
                sheet_name=sheet_name,
                dtype=str,  # Tüm değerleri string olarak oku
                na_filter=False  # NaN'ları boş string yap
            )
            
            logger.info(f"Okunan satır sayısı: {len(self.data)}")
            logger.info(f"Sütunlar: {list(self.data.columns)}")
            
            return self.data
            
        except Exception as e:
            logger.error(f"Excel okuma hatası: {str(e)}")
            raise FileOperationError(f"Excel dosyası okunamadı: {str(e)}")
    
    def get_sheet_names(self) -> List[str]:
        """Workbook'taki sheet isimlerini al"""
        if self.workbook:
            return list(self.workbook.sheetnames)
        return []
    
    def close(self):
        """Kaynakları temizle"""
        if self.workbook:
            self.workbook.close()


class ExcelWriter:
    """Excel dosya yazma sınıfı"""
    
    def __init__(self, output_path: Path):
        self.output_path = Path(output_path)
        self.workbook = None
        self.worksheets = {}
        
    def create_workbook(self) -> 'ExcelWriter':
        """Yeni workbook oluştur"""
        try:
            # Çakışma kontrolü
            final_path = handle_file_collision(self.output_path)
            if final_path != self.output_path:
                logger.info(f"Dosya çakışması tespit edildi. Yeni ad: {final_path}")
                self.output_path = final_path
            
            # Ana dizini oluştur
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Workbook oluştur
            self.workbook = xlsxwriter.Workbook(
                str(self.output_path), 
                {'strings_to_urls': False}  # URL dönüşümünü kapat
            )
            
            logger.info(f"Excel dosyası oluşturuluyor: {self.output_path}")
            return self
            
        except Exception as e:
            logger.error(f"Excel oluşturma hatası: {str(e)}")
            raise FileOperationError(f"Excel dosyası oluşturulamadı: {str(e)}")
    
    def add_worksheet(self, name: str) -> Any:
        """Yeni worksheet ekle"""
        if not self.workbook:
            raise FileOperationError("Workbook oluşturulmamış")
        
        # Sheet adını güvenli hale getir
        safe_name = safe_filename(name)[:31]  # Excel limit
        
        worksheet = self.workbook.add_worksheet(safe_name)
        self.worksheets[name] = worksheet
        
        logger.debug(f"Worksheet eklendi: {safe_name}")
        return worksheet
    
    def write_cikti_sheet(self, lines: List[Any]) -> None:
        """
        Çıktı sheet'ini yaz - YENİ FORMAT
        
        Kolonlar:
        A: Kart Listesi (text) - ANA KOLON
        B: Görsel Dosyası - ANA KOLON (Part 2 için boş)
        C: player_name - eşleştirme için
        D: series_name - eşleştirme için
        E: group - eşleştirme için
        F: denominator - eşleştirme için
        G: is_signed - eşleştirme için
        """
        worksheet = self.add_worksheet("Çıktı")
        
        # Format
        header_format = self.workbook.add_format({
            'bold': True,
            'bg_color': '#D9E1F2',
            'border': 1
        })
        
        # Headers - DOĞRU SIRALAMA
        headers = [
            "Kart Listesi",      # A - ana
            "Görsel Dosyası",    # B - ana (Part 2 dolduracak)
            "player_name",       # C - eşleştirme
            "series_name",       # D - eşleştirme
            "group",             # E - eşleştirme
            "denominator",       # F - eşleştirme
            "is_signed"          # G - eşleştirme
        ]
        
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Data - CardLine objelerinden al
        for i, line in enumerate(lines, 1):
            # line bir CardLine objesi mi yoksa string mi kontrol et
            if hasattr(line, 'text'):
                # CardLine objesi
                worksheet.write(i, 0, line.text)                              # A: Kart Listesi
                worksheet.write(i, 1, "")                                     # B: Görsel Dosyası (Part 2)
                worksheet.write(i, 2, line.player)                            # C: player_name
                worksheet.write(i, 3, line.series)                            # D: series_name
                worksheet.write(i, 4, line.group if line.group else "")       # E: group
                worksheet.write(i, 5, line.denominator)                       # F: denominator
                worksheet.write(i, 6, "Evet" if line.is_signed else "Hayır")  # G: is_signed
            else:
                # Geriye uyumluluk: string ise sadece A kolonuna yaz
                worksheet.write(i, 0, str(line))
        
        # Genişlik ayarla
        worksheet.set_column('A:A', 55)  # Kart Listesi
        worksheet.set_column('B:B', 55)  # Görsel Dosyası
        worksheet.set_column('C:C', 25)  # player_name
        worksheet.set_column('D:D', 20)  # series_name
        worksheet.set_column('E:E', 35)  # group
        worksheet.set_column('F:F', 12)  # denominator
        worksheet.set_column('G:G', 10)  # is_signed
        
        logger.info(f"Çıktı sheet'i yazıldı: {len(lines)} satır, 7 kolon")
    
    def write_ozet_sheet(self, summary_data: Dict[str, Any]) -> None:
        """Özet sheet'ini yaz"""
        worksheet = self.add_worksheet("Özet")
        
        # Format
        header_format = self.workbook.add_format({
            'bold': True,
            'bg_color': '#E2EFDA',
            'border': 1
        })
        
        bold_format = self.workbook.add_format({'bold': True})
        
        row = 0
        
        # Genel özet
        worksheet.write(row, 0, "GENEL ÖZET", header_format)
        row += 1
        
        worksheet.write(row, 0, "Toplam Kart Sayısı:", bold_format)
        worksheet.write(row, 1, summary_data.get('total_cards', 0))
        row += 1
        
        worksheet.write(row, 0, "Toplam Oyuncu Sayısı:", bold_format)
        worksheet.write(row, 1, summary_data.get('total_players', 0))
        row += 2
        
        # Variant özeti
        if 'variants' in summary_data:
            worksheet.write(row, 0, "VARIANT ÖZETI", header_format)
            row += 1
            
            for variant, counts in summary_data['variants'].items():
                worksheet.write(row, 0, f"{variant}:", bold_format)
                worksheet.write(row, 1, f"Normal: {counts.get('normal', 0)}")
                worksheet.write(row, 2, f"İmzalı: {counts.get('signed', 0)}")
                row += 1
        
        # Base özeti
        if 'base_summary' in summary_data:
            row += 1
            worksheet.write(row, 0, "BASE KARTLAR", header_format)
            row += 1
            
            for player, base_count in summary_data['base_summary'].items():
                worksheet.write(row, 0, player, bold_format)
                worksheet.write(row, 1, f"{base_count}X Base")
                row += 1
        
        # Genişlik ayarla
        worksheet.set_column('A:A', 25)
        worksheet.set_column('B:C', 15)
        
        logger.info("Özet sheet'i yazıldı")
    
    def write_hatalar_sheet(self, errors: List[Dict[str, Any]]) -> None:
        """Hatalar sheet'ini yaz"""
        worksheet = self.add_worksheet("Hatalar")
        
        # Format
        header_format = self.workbook.add_format({
            'bold': True,
            'bg_color': '#FFCCCB',
            'border': 1
        })
        
        error_format = self.workbook.add_format({
            'bg_color': '#FFE6E6'
        })
        
        # Headers
        headers = ["Satır", "Sütun", "Hata Türü", "Açıklama"]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Data
        for i, error in enumerate(errors, 1):
            worksheet.write(i, 0, error.get('row', ''), error_format)
            worksheet.write(i, 1, error.get('column', ''), error_format)
            worksheet.write(i, 2, error.get('type', ''), error_format)
            worksheet.write(i, 3, error.get('message', ''), error_format)
        
        # Genişlik ayarla
        worksheet.set_column('A:A', 10)
        worksheet.set_column('B:B', 15)
        worksheet.set_column('C:C', 20)
        worksheet.set_column('D:D', 50)
        
        logger.info(f"Hatalar sheet'i yazıldı: {len(errors)} hata")
    
    def write_uyarilar_sheet(self, warnings: List[Dict[str, Any]]) -> None:
        """Uyarılar sheet'ini yaz"""
        worksheet = self.add_worksheet("Uyarılar")
        
        # Format
        header_format = self.workbook.add_format({
            'bold': True,
            'bg_color': '#FFF2CC',
            'border': 1
        })
        
        warning_format = self.workbook.add_format({
            'bg_color': '#FFFACD'
        })
        
        # Headers
        headers = ["Satır", "Sütun", "Uyarı Türü", "Açıklama"]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Data
        for i, warning in enumerate(warnings, 1):
            worksheet.write(i, 0, warning.get('row', ''), warning_format)
            worksheet.write(i, 1, warning.get('column', ''), warning_format)
            worksheet.write(i, 2, warning.get('type', ''), warning_format)
            worksheet.write(i, 3, warning.get('message', ''), warning_format)
        
        # Genişlik ayarla
        worksheet.set_column('A:A', 10)
        worksheet.set_column('B:B', 15)
        worksheet.set_column('C:C', 20)
        worksheet.set_column('D:D', 50)
        
        logger.info(f"Uyarılar sheet'i yazıldı: {len(warnings)} uyarı")
    
    def write_ayarlar_sheet(self, config_data: Dict[str, Any]) -> None:
        """Ayarlar sheet'ini yaz"""
        worksheet = self.add_worksheet("Ayarlar")
        
        # Format
        header_format = self.workbook.add_format({
            'bold': True,
            'bg_color': '#F2F2F2',
            'border': 1
        })
        
        bold_format = self.workbook.add_format({'bold': True})
        
        row = 0
        
        # Program bilgileri
        worksheet.write(row, 0, "PROGRAM BİLGİLERİ", header_format)
        row += 1
        
        program_info = [
            ("Program Adı", PROGRAM_NAME),
            ("Versiyon", PROGRAM_VERSION),
            ("İşlem Tarihi", ist_timestamp()),
            ("Kullanıcı", get_current_user()),
            ("Timezone", "Europe/Istanbul")
        ]
        
        for key, value in program_info:
            worksheet.write(row, 0, key, bold_format)
            worksheet.write(row, 1, str(value))
            row += 1
        
        row += 1
        
        # İşlem ayarları
        worksheet.write(row, 0, "İŞLEM AYARLARI", header_format)
        row += 1
        
        for key, value in config_data.items():
            worksheet.write(row, 0, str(key), bold_format)
            worksheet.write(row, 1, str(value))
            row += 1
        
        # Genişlik ayarla
        worksheet.set_column('A:A', 25)
        worksheet.set_column('B:B', 40)
        
        logger.info("Ayarlar sheet'i yazıldı")
    
    def close(self):
        """Workbook'u kapat ve kaydet"""
        if self.workbook:
            self.workbook.close()
            logger.info(f"Excel dosyası kaydedildi: {self.output_path}")
        
        return self.output_path


def read_checklist_excel(file_path: Path) -> pd.DataFrame:
    """Checklist Excel dosyasını oku (convenience function)"""
    reader = ExcelReader(file_path)
    try:
        return reader.read_checklist()
    finally:
        reader.close()


def create_output_excel(
    output_path: Path,
    lines: List[Any],  # CardLine listesi veya string listesi
    summary_data: Dict[str, Any],
    errors: List[Dict[str, Any]],
    warnings: List[Dict[str, Any]],
    config_data: Dict[str, Any]
) -> Path:
    """Çıktı Excel dosyasını oluştur (convenience function)"""
    writer = ExcelWriter(output_path)
    
    try:
        writer.create_workbook()
        writer.write_cikti_sheet(lines)
        writer.write_ozet_sheet(summary_data)
        writer.write_hatalar_sheet(errors)
        writer.write_uyarilar_sheet(warnings)
        writer.write_ayarlar_sheet(config_data)
        
        return writer.close()
        
    except Exception as e:
        logger.error(f"Excel yazma hatası: {str(e)}")
        raise FileOperationError(f"Excel dosyası yazılamadı: {str(e)}")


if __name__ == "__main__":
    # Test amaçlı
    import logging
    logging.basicConfig(level=logging.INFO)
    
    # Test data
    test_lines = ["Test Kart 1", "Test Kart 2"]
    test_summary = {"total_cards": 2, "total_players": 1}
    test_errors = []
    test_warnings = []
    test_config = {"input_file": "test.xlsx"}
    
    output_path = Path("test_output.xlsx")
    result_path = create_output_excel(
        output_path, test_lines, test_summary, 
        test_errors, test_warnings, test_config
    )
    print(f"Test Excel oluşturuldu: {result_path}")